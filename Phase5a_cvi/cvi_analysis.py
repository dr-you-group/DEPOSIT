#!/usr/bin/env python3
"""
DEPOSIT checklist — Phase 5a: Content Validity Index (CVI) analysis.

Computes item- and scale-level content validity from an expert panel that
rated each checkpoint on three criteria:
  - Relevance    (4-point; 3-4 = endorsed)
  - Clarity      (4-point; 3-4 = endorsed)
  - Essentiality (E / U / N; E = essential, used for Lawshe's CVR)

Outputs:
  - I-CVI (item-level content validity index) for relevance and clarity
  - S-CVI/Ave and S-CVI/UA (scale-level, averaging and universal-agreement)
  - modified kappa (chance-corrected agreement; Polit, Beck & Owen 2007)
  - CVR (Lawshe 1975), reported descriptively
  - per-item CSV and per-stage summary

Input:  a directory of one .xlsx per rater, each with a sheet ('평정폼' by
        default) whose columns hold, per checkpoint row:
        No | Stage | Dimension | Sub-dimension | Checkpoint(EN) | Checkpoint(KR)
        | Relevance | Clarity | Essentiality | Comment
        Rater identity is taken from file order (rater_01, rater_02, ...), so
        no personal names appear in the code or output.

Usage:  python cvi_analysis.py --input ./data/exercise_a --out results_cvi.csv

References:
  Lynn (1986) Nurs Res 35:382-385.
  Lawshe (1975) Pers Psychol 28:563-575.
  Polit & Beck (2006) Res Nurs Health 29:489-497.
  Polit, Beck & Owen (2007) Res Nurs Health 30:459-467.
"""
import argparse, glob, os, re, math
import pandas as pd

SHEET = "평정폼"
COL = dict(no=1, stage=2, dim=3, en=5, rel=7, cla=8, ess=9, note=10)

def _num(v):
    if v is None: return None
    m = re.match(r"\s*([1-4])", str(v)); return int(m.group(1)) if m else None

def _ess(v):
    if v is None: return None
    m = re.match(r"\s*([EUN])", str(v).strip()); return m.group(1) if m else None

def load(input_dir):
    from openpyxl import load_workbook
    files = sorted(glob.glob(os.path.join(input_dir, "*.xlsx")))
    if not files:
        raise SystemExit(f"No .xlsx files found in {input_dir}")
    rows, meta = [], {}
    for i, f in enumerate(files, 1):
        rid = f"rater_{i:02d}"                     # anonymized by file order
        ws = load_workbook(f, data_only=True)[SHEET]
        stage = None
        for r in range(2, ws.max_row + 1):
            s = ws.cell(r, COL["stage"]).value
            if s: stage = s
            item = ws.cell(r, COL["en"]).value
            if not item: continue
            no = ws.cell(r, COL["no"]).value
            no = int(str(no).strip()) if no is not None else None
            meta[no] = dict(no=no, stage=stage, dim=ws.cell(r, COL["dim"]).value, item=str(item))
            rows.append(dict(no=no, rater=rid,
                             rel=_num(ws.cell(r, COL["rel"]).value),
                             cla=_num(ws.cell(r, COL["cla"]).value),
                             ess=_ess(ws.cell(r, COL["ess"]).value)))
    return pd.DataFrame(rows), pd.DataFrame(meta).T.sort_values("no")

def cvi_table(d, col):
    out = []
    for no, g in d.groupby("no"):
        v = g[col].dropna(); n = len(v); a = int((v >= 3).sum())
        icvi = a / n if n else float("nan")
        pc = (math.comb(n, a)) * (0.5 ** n) if n else float("nan")   # chance agreement
        kappa = (icvi - pc) / (1 - pc) if n and pc < 1 else float("nan")
        out.append(dict(no=no, n=n, agree=a, cvi=icvi, kappa=kappa))
    return pd.DataFrame(out).set_index("no")

def cvr_table(d):
    out = []
    for no, g in d.groupby("no"):
        v = g.ess.dropna(); n = len(v); ne = int((v == "E").sum())
        out.append(dict(no=no, n=n, n_E=ne, cvr=(ne - n / 2) / (n / 2) if n else float("nan")))
    return pd.DataFrame(out).set_index("no")

def kcat(k):
    if k != k: return None
    return "excellent" if k > 0.74 else "good" if k > 0.59 else "fair" if k > 0.39 else "poor"

def main():
    ap = argparse.ArgumentParser(description="DEPOSIT Phase 5a — CVI analysis")
    ap.add_argument("--input", required=True, help="directory of one .xlsx per rater")
    ap.add_argument("--out", default="results_cvi.csv", help="per-item output CSV")
    a = ap.parse_args()

    d, M = load(a.input)
    n_raters = d.rater.nunique()
    print(f"raters={n_raters}  items={d.no.nunique()}")

    rel, cla, cvr = cvi_table(d, "rel"), cvi_table(d, "cla"), cvr_table(d)
    res = (M.set_index("no")
             .join(rel.add_prefix("rel_")).join(cla.add_prefix("cla_")).join(cvr.add_prefix("ess_")))
    res["stage"] = res["stage"].astype(str)

    print("\n=== Scale level ===")
    for name, t in [("Relevance", rel), ("Clarity", cla)]:
        print(f"  {name}: S-CVI/Ave={t.cvi.mean():.3f}  S-CVI/UA={(t.cvi==1).mean():.3f}")
    print(f"  Mean CVR={cvr.cvr.mean():.3f}")

    print("\n=== By stage ===")
    for st, g in res.groupby("stage"):
        print(f"  {st}: n={len(g)}  relAve={g.rel_cvi.mean():.3f}  claAve={g.cla_cvi.mean():.3f}  meanCVR={g.ess_cvr.mean():.3f}")

    res["rel_kappa_cat"] = res.rel_kappa.map(kcat)
    print("\n=== modified kappa (relevance) ===")
    print(res.rel_kappa_cat.value_counts().to_string())

    print(f"\n=== I-CVI(relevance) < 0.78 ===")
    for no, r in res[res.rel_cvi < 0.77].iterrows():
        print(f"  #{no} [{r.stage}] I-CVI={r.rel_cvi:.2f}")

    res.to_csv(a.out, encoding="utf-8-sig")
    print(f"\nsaved: {a.out}")

if __name__ == "__main__":
    main()
