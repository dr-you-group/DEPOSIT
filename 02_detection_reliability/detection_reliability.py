#!/usr/bin/env python3
"""
DEPOSIT checklist — Phase 5b: detection sensitivity + inter-rater reliability.

Raters independently apply the Stage 2-4 checkpoints to synthetic submission
files that contain known, seeded defects, judging each checkpoint per file as
pass / fail / not-applicable and recording free-text evidence for failures.

Outputs:
  (A) Detection sensitivity vs. the seeded-defect answer key
      (union / majority / unanimous; by stage, data type, difficulty).
  (B) Inter-rater agreement on the pass/fail/NA judgments:
      Fleiss' kappa and Gwet's AC1 with bootstrap 95% CIs.
      AC1 is reported because a high prevalence of NA judgments deflates kappa.

Inputs:
  --responses  directory of one .xlsx per rater (rater identity taken from
               file order: rater_01, rater_02, ...). Each file has a sheet
               ('평정폼') with, per row:
               No | Stage | Dimension | Sub-dim | Checkpoint(EN) | Checkpoint(KR)
               | judgment_file1 | judgment_file2 | judgment_file3 | evidence
  --answer-key CSV of seeded defects with columns:
               defect_id, file, stage, difficulty, is_trap, description,
               signature   (a regex matched against a rater's evidence text)
  --files      comma-separated data-type labels matching the three judgment
               columns, in order (default: file1,file2,file3).

All input data are SYNTHETIC; the "identifiers" in the answer key are
fabricated values seeded as defects, not real personal data.

Usage:
  python detection_reliability.py \
      --responses ./data/exercise_b \
      --answer-key ./data/answer_key.csv \
      --files imaging,genomic,clinical

References:
  Gwet (2008) Br J Math Stat Psychol 61:29-48.
  Kottner et al. (2011) J Clin Epidemiol 64:96-106 (GRRAS).
"""
import argparse, glob, os, re
import numpy as np, pandas as pd

SHEET = "平정폼".replace("平","평")  # '평정폼'
CATS = ["예", "아니오", "해당없음"]   # pass / fail / not-applicable
JUDG_COLS = (7, 8, 9)                 # three per-file judgment columns
EN_COL, NO_COL, STAGE_COL, EV_COL = 5, 1, 2, 10

def load_responses(resp_dir, file_labels):
    from openpyxl import load_workbook
    files = sorted(glob.glob(os.path.join(resp_dir, "*.xlsx")))
    if not files:
        raise SystemExit(f"No .xlsx files in {resp_dir}")
    J, EV, items, raters = {}, {}, {}, []
    for i, f in enumerate(files, 1):
        rid = f"rater_{i:02d}"; raters.append(rid)
        ws = load_workbook(f, data_only=True)[SHEET]
        stage = None
        for r in range(2, ws.max_row + 1):
            s = ws.cell(r, STAGE_COL).value
            if s: stage = s
            if not ws.cell(r, EN_COL).value: continue
            no = int(str(ws.cell(r, NO_COL).value).strip())
            items[no] = stage
            for lab, c in zip(file_labels, JUDG_COLS):
                v = ws.cell(r, c).value
                J[(rid, no, lab)] = str(v).strip() if v else None
            EV[(rid, no)] = str(ws.cell(r, EV_COL).value or "")
    return J, EV, items, raters, file_labels

def detection(EV, raters, ak):
    corpus = {rid: " ".join(EV.get((rid, no), "") for no in {k[1] for k in EV}).lower()
              for rid in raters}
    real = ak[ak.is_trap.isna()] if "is_trap" in ak else ak
    rows = []
    for _, d in real.iterrows():
        det = {rid: bool(re.search(str(d.signature), corpus[rid])) for rid in raters}
        rows.append(dict(defect_id=d.defect_id, file=d.file, stage=d.stage,
                         difficulty=d.get("difficulty", ""),
                         n_detected=sum(det.values())))
    return pd.DataFrame(rows)

def fleiss_ac1(mat):
    """mat: (units x categories) counts per unit. Returns (Fleiss k, Gwet AC1)."""
    n_u, q = mat.shape; n_r = mat.sum(1)[0]
    p_j = mat.sum(0) / (n_u * n_r)
    P_i = ((mat ** 2).sum(1) - n_r) / (n_r * (n_r - 1))
    Pbar = P_i.mean()
    Pe_f = (p_j ** 2).sum()
    Pe_g = (1 / (q - 1)) * sum(p * (1 - p) for p in p_j)
    k = (Pbar - Pe_f) / (1 - Pe_f)
    ac1 = (Pbar - Pe_g) / (1 - Pe_g)
    return k, ac1

def build_matrix(J, items, raters, labels, stage=None):
    u = []
    for no in items:
        if stage is not None and not str(items[no]).endswith(str(stage)): continue
        for lab in labels:
            v = [J[(rid, no, lab)] for rid in raters]
            v = [x for x in v if x in CATS]
            if len(v) == len(raters):
                u.append([v.count(c) for c in CATS])
    return np.array(u)

def boot_ci(mat, B=5000, seed=0):
    rng = np.random.default_rng(seed); ks, as_ = [], []
    for _ in range(B):
        idx = rng.integers(0, mat.shape[0], mat.shape[0])
        k, a = fleiss_ac1(mat[idx])
        if np.isfinite(k): ks.append(k)
        if np.isfinite(a): as_.append(a)
    return np.percentile(ks, [2.5, 97.5]), np.percentile(as_, [2.5, 97.5])

def main():
    ap = argparse.ArgumentParser(description="DEPOSIT Phase 5b — detection + reliability")
    ap.add_argument("--responses", required=True)
    ap.add_argument("--answer-key", required=True)
    ap.add_argument("--files", default="KBI,KRA,GeNA")
    ap.add_argument("--out", default="results_detection.csv")
    a = ap.parse_args()
    labels = [x.strip() for x in a.files.split(",")]

    J, EV, items, raters, labels = load_responses(a.responses, labels)
    ak = pd.read_csv(a.answer_key)
    print(f"raters={len(raters)} items={len(items)} files={labels}")

    # ---- (A) detection ----
    D = detection(EV, raters, ak)
    tot = len(D)
    print(f"\n=== Detection (n={tot} seeded defects) ===")
    print(f"  union (>=1): {(D.n_detected>=1).sum()}/{tot} = {(D.n_detected>=1).mean():.1%}")
    print(f"  majority(>=2): {(D.n_detected>=2).sum()}/{tot}")
    print(f"  unanimous  : {(D.n_detected==len(raters)).sum()}/{tot}")
    for key in ("stage", "file", "difficulty"):
        if key in D:
            print(f"  by {key}: " + ", ".join(f"{k} {(g.n_detected>=1).sum()}/{len(g)}"
                                              for k, g in D.groupby(key)))
    D.to_csv(a.out, index=False, encoding="utf-8-sig")

    # ---- (B) reliability ----
    M = build_matrix(J, items, raters, labels)
    na_prev = M[:, CATS.index("해당없음")].sum() / M.sum()
    k, ac1 = fleiss_ac1(M); (kl, kh), (al, ah) = boot_ci(M)
    print(f"\n=== Inter-rater agreement (n={M.shape[0]} judgments; NA={na_prev:.1%}) ===")
    print(f"  Fleiss' kappa = {k:.3f} [{kl:.3f}, {kh:.3f}]")
    print(f"  Gwet's AC1    = {ac1:.3f} [{al:.3f}, {ah:.3f}]")
    for st in sorted({str(v)[-1] for v in items.values()}):
        Ms = build_matrix(J, items, raters, labels, stage=st)
        if len(Ms) < 3: continue
        k2, a2 = fleiss_ac1(Ms)
        print(f"  Stage {st} (n={Ms.shape[0]}): kappa={k2:.3f}, AC1={a2:.3f}")
    print(f"\nsaved: {a.out}")

if __name__ == "__main__":
    main()
