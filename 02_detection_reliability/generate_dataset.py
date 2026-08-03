#!/usr/bin/env python3
"""
DEPOSIT Exercise 2 — LINKED multimodal evaluation dataset.
30 subjects (SUBJ001-030) shared across KBI (imaging), KRA (genomics), GeNA (clinical).
27 seeded defects, Stage-2 emphasis: Stage2=17, Stage3=5, Stage4=5.
Files: KBI 11 / KRA 11 / GeNA 5. Cross-modal re-identification traps included
(not too subtle). Fills the ACTUAL uploaded templates.
"""
import os
from openpyxl import load_workbook

FIX="/home/claude/fillwork"
OUT="/home/claude/ex2/dataset"; AK="/home/claude/ex2/answer_key"
os.makedirs(OUT,exist_ok=True); os.makedirs(AK,exist_ok=True)

N=30
SUBJ=[f"SUBJ{i:03d}" for i in range(1,N+1)]      # shared pseudonymous IDs
SEX=["M" if i%2==0 else "F" for i in range(N)]
AGE=[55+(i%25) for i in range(N)]                # 55-79 plausible

answer=[]
def log(did,file,stage,item,field,xlrow,xlcol,dtype,diff,desc,value,trap=False):
    answer.append(dict(defect_id=did,file=file,stage=stage,checklist_item_id=item,field=field,
        excel_row=xlrow,excel_col=xlcol,defect_type=dtype,difficulty=diff,description=desc,
        seeded_value=value,is_trap=("TRAP-normal" if trap else "")))
def col(n):
    s=""
    while n>0: n,r=divmod(n-1,26); s=chr(65+r)+s
    return s

# ============================ KBI (KBI_Data, start row 5, 50 cols) ============================
KS=5
def kbi_clean(i):
    s=SUBJ[i]
    return [f"Brain MR image {s}","",f"{s}.dcm","Homo sapiens","","Brain","",
            s,SEX[i],AGE[i],"","","type 2 diabetes mellitus","cohort","",
            "T2DM brain MRI","brain // T2DM // RWE","2024-03-14","Severance Hospital","T1","",
            "SIEMENS","Prisma","","","syngo MR E11","raw","1.8.0","3","HeadCoil-32ch",
            "","","","","","Gradient Echo","","","","","","","","","","","","j-","",""]
kbi=[kbi_clean(i) for i in range(N)]
# ---- Stage 2 (7) — de-identification, KBI ----
kbi[0][7]="Kim Minjun";                 log("KBI-01","KBI",2,"B1005","Patient ID",KS+0,8,"PHI (name)","obvious","Patient ID에 가명 대신 실명 잔존 (SUBJ001의 실제 신원)","Kim Minjun")
kbi[1][7]="SUBJ002 / 880201-2xxxxxx";   log("KBI-02","KBI",2,"B1006","Patient ID",KS+1,8,"PHI (RRN)","obvious","Patient ID에 주민등록번호 병기","880201-2xxxxxx")
kbi[2][2]="KimMinjun_20240314.dcm";     log("KBI-03","KBI",2,"B1018","Raw data",KS+2,3,"PHI in file name","moderate","실데이터 파일명에 실명 포함","KimMinjun_20240314.dcm")
kbi[3][23]="SN-MR-118277";              log("KBI-04","KBI",2,"B1017","Device serial number",KS+3,24,"Device identifier","moderate","장치 일련번호 잔존","SN-MR-118277")
kbi[4][17]="2024-03-14 (홍길동)";        log("KBI-05","KBI",2,"B1005","Exp_date",KS+4,18,"PHI (name) in field","moderate","Exp_date 칸에 환자명 혼입","2024-03-14 (홍길동)")
kbi[5][9]=93;                            log("KBI-06","KBI",2,"B1007","Patient age",KS+5,10,"Age 90+ not generalized","moderate","90세 이상 미상한처리(연령 특이값)","93")
kbi[6][19]="T1 (Seoul, Gangnam-gu)";    log("KBI-07","KBI",2,"B1008","Data type",KS+6,20,"Geographic identifier leaked","moderate","값 칸에 시군구 지역정보 혼입","Seoul, Gangnam-gu")
# ---- Stage 3 (2) ----
kbi[7][35]="";                          log("KBI-08","KBI",3,"C2004","Pulse sequence type",KS+7,36,"Mandatory missing","obvious","필수항목 Pulse sequence type 누락","(blank)")
kbi[9]=list(kbi[8]);                     log("KBI-09","KBI",3,"C3009","(entire row)",KS+9,0,"Duplicate record","moderate","앞 행과 완전 중복 (SUBJ009 사본)","duplicate of row 13")
# ---- Stage 4 (2) ----
kbi[10][19]="MRI-scan";                 log("KBI-10","KBI",4,"D1002","Data type",KS+10,20,"Outside controlled vocab","moderate","Data type 통제어휘(T1/T2/fMRI…) 위반","MRI-scan")
kbi[11][17]="03/15/2024";               log("KBI-11","KBI",4,"D4001","Exp_date",KS+11,18,"Non-ISO date","moderate","날짜 포맷 비표준(YYYY-MM-DD 아님)","03/15/2024")
# ---- TRAP (normal, must NOT be flagged) ----
kbi[12][7]="SUBJ013";                   log("KBI-T1","KBI",2,"(none)","Patient ID",KS+12,8,"TRAP: valid pseudonym","-","함정: 정상 가명 ID (개인정보 아님)","SUBJ013",trap=True)

# ============================ KRA (KRA_Metadata, start row 10, 18 cols) ============================
RS=10
def kra_clean(i):
    s=SUBJ[i]
    return [s,f"lib_{s}",f"WGS of {s}","WGS","GENOMIC","RANDOM","PAIRED","ILLUMINA",
            "Illumina NovaSeq 6000","TruSeq DNA PCR-Free","fastq",f"{s}_R1.fastq.gz","","","","","",""]
kra=[kra_clean(i) for i in range(N)]
# ---- Stage 2 (7) ----
kra[0][0]="880101-1234567";             log("KRA-01","KRA",2,"B1006","sample name",RS+0,1,"PHI (RRN)","obvious","sample name에 주민등록번호","880101-1234567")
kra[1][2]="WGS of patient Kim Minjun";  log("KRA-02","KRA",2,"B1005","title",RS+1,3,"PHI (name)","moderate","title에 실명 포함","...Kim Minjun")
kra[2][11]="ParkJiSoo_G003_R1.fastq.gz";log("KRA-03","KRA",2,"B3010","raw file",RS+2,12,"PHI in file name","moderate","원시데이터 파일명에 실명","ParkJiSoo_...fastq.gz")
kra[3][1]="lib_홍길동";                  log("KRA-04","KRA",2,"B1005","library name",RS+3,2,"PHI (name)","moderate","library name에 실명 혼입","lib_홍길동")
# cross-modal re-ID: KRA sample name = SUBJ005 pseudonym, but a real name is exposed elsewhere for same subject
kra[4][2]="WGS of SUBJ005 (환자: 이수진)"; log("KRA-05","KRA",2,"B1005","title",RS+4,3,"Cross-modal re-ID","moderate","가명(SUBJ005)에 실명 병기 → 타 파일과 대조 시 재식별","SUBJ005 (이수진)")
kra[5][9]="protocol; contact hjkim@yuhs.ac"; log("KRA-06","KRA",2,"B1015","library construction protocol",RS+5,10,"PHI (email)","moderate","프로토콜 설명에 담당자 이메일 노출","hjkim@yuhs.ac")
kra[6][0]="SUBJ007_010-2345-6789";      log("KRA-07","KRA",2,"B1013","sample name",RS+6,1,"PHI (phone)","obvious","sample name에 휴대폰번호 결합","010-2345-6789")
# ---- Stage 3 (2) ----
kra[7][7]="";                           log("KRA-08","KRA",3,"C2004","platform",RS+7,8,"Mandatory missing","obvious","필수항목 platform 누락","(blank)")
kra[9]=list(kra[8]);                     log("KRA-09","KRA",3,"C3009","(entire row)",RS+9,0,"Duplicate record","moderate","앞 행과 완전 중복 (SUBJ009 사본)","duplicate of row 18")
# ---- Stage 4 (2) ----
kra[10][3]="WholeGenome";               log("KRA-10","KRA",4,"D1002","strategy",RS+10,4,"Outside controlled vocab","moderate","strategy 통제어휘 위반(WGS 아님)","WholeGenome")
kra[11][7]="illumina";                  log("KRA-11","KRA",4,"D4003","platform",RS+11,8,"Inconsistent case","subtle","platform 대소문자 불일치","illumina")
# ---- TRAP ----
kra[12][0]="SUBJ013";                   log("KRA-T1","KRA",2,"(none)","sample name",RS+12,1,"TRAP: valid pseudonym","-","함정: 정상 가명 ID","SUBJ013",trap=True)

# ============================ GeNA (GeNA, start row 5, 12 cols) ============================
GS=5
def gena_clean(i):
    s=SUBJ[i]
    return [s,"기타(보건정보)","","Y","De-identified T2DM clinical dataset","diabetes; CKD; RWE",
            "","clinical CSV file",f"clinical_{s}.csv","","","2026-12-31"]
gena=[gena_clean(i) for i in range(N)]
# ---- Stage 2 (3) ----
gena[0][4]="김민준 환자 임상데이터 (880101-1234567)"; log("GeNA-01","GeNA",2,"B1006","제목",GS+0,5,"PHI (name/RRN)","obvious","제목에 실명·주민번호","김민준 (880101-1234567)")
gena[1][6]="환자 연락처 010-2345-6789";  log("GeNA-02","GeNA",2,"B1013","추가 설명",GS+1,7,"PHI (phone)","obvious","추가설명에 전화번호","010-2345-6789")
gena[2][8]="ParkJiSoo_clinical.csv";     log("GeNA-03","GeNA",2,"B1018","파일",GS+2,9,"PHI in file name","moderate","파일명에 실명 포함 → KRA와 동일 인물 재식별","ParkJiSoo_clinical.csv")
# ---- Stage 3 (1) ----
gena[3][1]="";                           log("GeNA-04","GeNA",3,"C2004","범주",GS+3,2,"Mandatory missing","obvious","필수항목 범주 누락","(blank)")
# ---- Stage 4 (1) ----
gena[4][11]="2026/12/31";                log("GeNA-05","GeNA",4,"D4001","공개 날짜",GS+4,12,"Non-ISO date","moderate","공개날짜 포맷 비표준","2026/12/31")
# ---- TRAP ----
gena[5][0]="SUBJ006";                    log("GeNA-T1","GeNA",2,"(none)","샘플명",GS+5,1,"TRAP: valid pseudonym","-","함정: 정상 가명 ID","SUBJ006",trap=True)

# ============================ fill real templates ============================
def fill(fixed,out,sheet,rows,start):
    wb=load_workbook(fixed); ws=wb[sheet]
    for i,row in enumerate(rows):
        for j,v in enumerate(row):
            ws.cell(row=start+i,column=j+1,value=(v if v!="" else None))
    wb.save(out)
fill(f"{FIX}/KBI_MR_metadata_fixed.xlsx",f"{OUT}/KBI_MR_metadata.xlsx","KBI_Data",kbi,KS)
fill(f"{FIX}/KRA_metadata_fixed.xlsx",  f"{OUT}/KRA_metadata.xlsx","KRA_Metadata",kra,RS)
fill(f"{FIX}/GeNA_metadata_fixed.xlsx", f"{OUT}/GeNA_metadata.xlsx","GeNA",gena,GS)

import pandas as pd
ad=pd.DataFrame(answer)
ad["excel_cell"]=[f"{col(c)}{r}" if c>0 else f"row {r}" for c,r in zip(ad.excel_col,ad.excel_row)]
ad=ad[["defect_id","file","stage","checklist_item_id","field","excel_cell","defect_type","difficulty","is_trap","description","seeded_value"]]
ad.to_csv(f"{AK}/seeded_defects_answer_key.csv",index=False)
real=ad[ad.is_trap==""]
print("subjects:",N," | rows: KBI",N,"KRA",N,"GeNA",N)
print("REAL defects:",len(real)," (+ traps:",len(ad)-len(real),")")
print("by file:",real.groupby('file').size().to_dict())
print("by stage:",real.groupby('stage').size().to_dict())
